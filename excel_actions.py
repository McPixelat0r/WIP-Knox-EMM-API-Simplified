from openpyxl import load_workbook
import easygui
from re import findall


# Gets the path of the Excel file to read and create instance of workbook
def createWorkbook():
    excel_file_path = easygui.fileopenbox()
    wb = load_workbook(filename=excel_file_path)
    ws = wb['Sheet1']
    return ws


# Get list of tab numbers
def getTabNums():
    num_list = []
    worksheet = createWorkbook()
    count = 0
    for i in worksheet['A']:
        cell_val = i.value
        if type(cell_val) is str:
            num_list.append([int(s) for s in findall(r'\b\d+\b', cell_val)][0])
            count += 1

        elif type(cell_val) is int:
            num_list.append(cell_val)
            count += 1
    return num_list


def getTabIMEIs():
    imei_list = []
    worksheet = createWorkbook()
    count = 0
    for i in worksheet['A']:
        cell_val = str(i.value)
        imei_list.append(cell_val)
        count += 1
    return imei_list


def getTabIds():
    num_list = []
    worksheet = createWorkbook()
    count = 0
    for i in worksheet['A']:
        cell_val = i.value
        if len(cell_val) > 1:
            num_list.append(cell_val)
    return num_list
